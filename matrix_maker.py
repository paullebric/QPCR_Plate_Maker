import numpy as np
import openpyxl
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter
from datetime import datetime
from math import ceil, floor
path="C:\\Users\\paulo\\OneDrive\\Bureau\\"

def main_input(char):
    échantillon = [input(char)]
    while True :
        échantillon += [input()]
        if échantillon[-1] == "":
            échantillon.pop(-1)
            break
    return échantillon

def apply_method(S,T,mode):
    fill=[]
    for s in S :
        for t in T :
            for i in range(mode):
                fill +=[s+" + "+t]
    return fill

def simple_plate(plate):
    for x in range(9) :
        for y in range(13):
            left = plate[x][0]
            top = plate[0][y]
            if left and top !="": plate[x][y] = left + " " + top
    for x in range(1,9):
        plate[x][0]=get_column_letter(x)
    for y in range(1,13):
        plate[0][y]=y
        
    return plate

def bordure(side):
    border_style = Border(
    left=Side(style=side["left"], color="000000") if 'left' in side else None,
    right=Side(style=side["right"], color="000000") if 'right' in side else None,
    top=Side(style=side["top"], color="000000") if 'top' in side else None,
    bottom=Side(style=side["bottom"], color="000000") if 'bottom'in side else None,
    )
    return border_style

def ecrire_matrice_excel(matrice,Sample,Target,mode):
    couleurs_fond = [
    "DDEBF7", 
    "FCE4D6", 
    "FFF2CC", 
    "E2EFDA", 
    "D6DCE4",  
    "C6F7D1",  
    "D0BFE3",  
    "FFA3A3",  
    "D6FAFF",  
    "F5FFD6",  
    "FFD6FD", 
    "C2ACAC"  
    ]

    couleurs_texte = [
    "D98719",  # Dark Red
    "FF1CAE",  # Dark Slate Gray
    "FF2400",  # Dark Gray
    "0000FF",  # Indigo
    "00EA75",  # Cadet Blue
    "238E23",  # Dark Olive Green
    "FF7F00",  # Slate Gray
    "4F2F4F",  # Sea Green
    "CFB53B",  # Slate Blue
    "2F4F4F",  # Saddle Brown
    "7093DB",  # Purple
    "5F9F9F"   # Dark Slate Blue
    ]

    nom_fichier = "QPCR_plate_"+str(datetime.now().date())+".xlsx"
    wb = openpyxl.Workbook()
    sheet = wb.active

    max_length =0
    for x in range(np.shape(matrice)[0]):
            for y in range(np.shape(matrice)[1]):
                if len(matrice[x][y]) > max_length:
                    max_length = len(matrice[x][y])
    
    for x in range(42):
        sheet.column_dimensions[get_column_letter(x+1)].width = max_length
        
    for i in range(matrice.shape[0]):  # Pour chaque ligne
        for j in range(matrice.shape[1]):  # Pour chaque colonne
            sheet.cell(row=i+1, column=j+1, value=matrice[i, j])
            currentCell = sheet[f"{get_column_letter(j+1)}{i+1}"] #or currentCell = ws['A1']
            currentCell.alignment = Alignment(horizontal='center')
             #or currentCell = ws['A1']
            # Insérer la valeur dans Excel
            # sheet[f"{get_column_letter(j+1)}{i+1}"].border = bordure(('left','right','top','bottom'),'thin')

    for x in range(len(Sample)) :
        couleur = couleurs_fond[x % len(couleurs_fond)]  # Assurer que les couleurs sont cyclées si plus d'échantillons que de couleurs
        fill = PatternFill(start_color=couleur, end_color=couleur, fill_type="solid")
        formula = f'ISNUMBER(SEARCH("{Sample[x]}", A1))'
        rule = FormulaRule(formula=[formula], fill=fill)
        sheet.conditional_formatting.add('A1:BZ10', rule)

    for x in range(len(Target)) :
        couleur_texte = couleurs_texte[x % len(couleurs_texte)]  # Assurer que les couleurs sont cyclées si plus d'échantillons que de couleurs
        font = Font(color=couleur_texte)
        formula = f'ISNUMBER(SEARCH("{Target[x]}", A1))'
        rule = FormulaRule(formula=[formula], font=font)
        sheet.conditional_formatting.add('A1:BZ10', rule)
    
    nb_plate = floor(matrice.shape[0]*matrice.shape[1]/96)
    count=0
    for tour in range(nb_plate):
        for x in range(count+0,count+13):
            for y in range(1,10):
                sheet[f"{get_column_letter(x+1)}{y}"].border = bordure({'left':'thick' if x-count==0 else 'thin',
                                                                        'right':'thick'if x-count==12 else 'thin',
                                                                        'top':'thick'if y==1 else 'thin',
                                                                        'bottom':'thick'if y==9 else 'thin'})
        count += 15
    wb.save(path+nom_fichier)
    print(f"Fichier {nom_fichier} créé avec succès!")

def plate_matrixer(S,T,mode):
    plate = np.zeros((9,13),dtype='U60')
    if len(S)<=8 and len(T)*mode<=12:
        left = S ;top = T
    elif len(T)<8 and len(S)*mode<=12:
        left = T; top = S
    else :
        return None
    for x in range(len(left)) :
        plate[x+1][0] = left[x]
    ntop=[]
    for x in top :
        for y in range(mode) : ntop += [x]
    for y in range(len(ntop)) : plate[0][y+1] = ntop[y]

    return simple_plate(plate)

def complex_plate_matrixer(Fill):
    nb_plate = ceil(len(Fill)/96)
    plate = np.zeros((9,13*nb_plate+2*nb_plate-1),dtype="U60")
    count=0
    for tour in range(nb_plate):
        for x in range(count+0,count+13):
            if x-count==0 :
                left =("A","B","C","D","E","F","G","H")
                for name,y in zip(left,range(1,9)):
                    plate[y][x]=name
            else :
                plate[0][x]=str(x-count)
        for y in range(1,9):
            for x in range(count+1,count+13):
                if len(Fill)>0:
                    plate[y][x]=Fill[0]
                    Fill.pop(0)
        count += 15
    return plate
    

def main():
    Sample= main_input("Noms des écchantillons/controles, Si tout les échantillons entrés clic enter :\n")
    Target= main_input("Noms des ammorces/targets, Si tout les targets entrés clic enter :\n")
    mode = int(input("Simplicat [1], Duplicat [2], Triplicat [3]"))
    Fillers = apply_method(Sample,Target,mode)
    plate = plate_matrixer(Sample,Target,mode)
    if plate is None:
        plate = complex_plate_matrixer(Fillers)
    return ecrire_matrice_excel(plate,Sample,Target,2)
    
main()
#plate_matrixer(["S12Allprep","S14Allprep","S12RNeasy","S14RNeasy"],["423","669","16-1","103","16-5","223"],3)